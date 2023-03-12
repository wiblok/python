from openpyxl import load_workbook

# Excelファイルの読み込み
wb = load_workbook('input.xlsx')

# シートの取得
ws = wb.worksheets[0]

# セルの取得
cell = ws.cell(row=1, column=1)

# セルに値を設定
cell.value = 'Hello, world!'

# Excelファイルの保存
wb.save('output.xlsx')